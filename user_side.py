import sqlite3
from datetime import datetime, date
from tkinter import *
from tkinter import messagebox

import cv2
import gspread
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials

import Style_font

# getting data from Database
con = sqlite3.connect('DataBase/client_detail.db')
cur = con.cursor()

recognizer = cv2.face.LBPHFaceRecognizer_create()
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);
# for spreadsheet
scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name('testSheets-9c1b98942881.json', scope)
client = gspread.authorize(creds)
sheet = client.open('fullbuster').sheet1


class user_():
    def __init__(self, root):
        self.root = root
        style = Style_font.Font_Style()
        label = Label(self.root, text="enter your ID", font=style.Label_Font)
        label.pack()

        self.entry = Entry(self.root)
        self.entry.pack()
        button = Button(self.root, text='Done', command=self.comform)
        button.pack()

    def comform(self):
        if str(self.entry.get() != ""):
            self.main_()
        else:
            messagebox.showerror('Error', 'Plesae enter your ID first')

    def main_(self):
        self.client_id = int(self.entry.get())
        person = cur.execute(f"select * from 'client_detail' where ID='{self.client_id}'")
        con.commit()
        for per in person:
            print(per)
        self.client_name = per[1]
        client_comform = 0
        no_of_loops = 0
        recognizer.read(f'trainer/trainer{self.client_id}.yml')

        cam = cv2.VideoCapture(0)

        while True:
            no_of_loops += 1

            ret, img = cam.read()

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            faces = faceCascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

            for (x, y, w, h) in faces:

                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

                id, confidence = recognizer.predict(gray[y:y + h, x:x + w])

                # Check if confidence is less them 100 ==> "0" is perfect match
                if (confidence < 100):
                    client_comform += 1
                else:
                    pass
            if cv2.waitKey(1) == 13 or no_of_loops == 10:  # Press 'ESC' for exiting video
                break

        if client_comform > 7:
            self.date_1 = str(date.today())
            self.time_1 = str(datetime.now().time())[:8]
            self.update_data()
        else:
            messagebox.showerror('Error', f'Client not matched ')
        # Do a bit of cleanup
        cam.release()
        cv2.destroyAllWindows()

    def update_data(self):
        insertrow = [self.client_id, self.client_name, self.time_1, self.date_1]
        sheet.insert_row(insertrow, 2)
        self.Local_attendance()
        messagebox.showinfo('Success', f'client comformed name = {self.client_name} " at " {self.time_1}')

    def Local_attendance(self):
        file = f'Student_base_attendance/{self.client_id}.xlsx'
        new_row = [self.client_id, self.client_name, self.time_1, self.date_1]

        wb = openpyxl.load_workbook(filename=file)
        ws = wb['Sheet1']  # Older method was  .get_sheet_by_name('Sheet1')
        row = ws.max_row + 1
        for col, entry in enumerate(new_row, start=1):
            ws.cell(row=row, column=col, value=entry)

        wb.save(file)


if __name__ == '__main__':
    root = Tk()
    root.geometry('400x300')
    root.resizable(False, False)
    atul = user_(root)
    root.mainloop()
